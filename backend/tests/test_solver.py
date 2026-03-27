from __future__ import annotations

from copy import deepcopy
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.exporters.csv_exporter import export_csv
from app.exporters.xlsx_exporter import export_xlsx
from app.main import app
from app.models.schemas import ScheduleRequest
from app.solver.engine import solve_schedule

client = TestClient(app)


def base_payload() -> dict:
    return {
        "month": 4,
        "year": 2026,
        "norm_hours": 168,
        "shifts": [
            {"id": "A2", "type": "day", "start_time": "07:00", "end_time": "19:00", "duration_hours": 12},
            {"id": "A3", "type": "night", "start_time": "19:00", "end_time": "07:00", "duration_hours": 12}
        ],
        "employees": [
            {"id": "e1", "name": "Ivan", "role": "Operator", "start_day": 1, "first_shift": "A2", "is_cover": False},
            {"id": "e2", "name": "Maria", "role": "Operator", "start_day": 1, "first_shift": "A3", "is_cover": False},
            {"id": "e3", "name": "Niki", "role": "Operator", "start_day": 2, "first_shift": "A2", "is_cover": False},
            {"id": "e4", "name": "Raya", "role": "Operator", "start_day": 3, "first_shift": "A3", "is_cover": False},
            {"id": "e5", "name": "Mila", "role": "Operator", "start_day": 4, "first_shift": "A2", "is_cover": False},
            {"id": "c1", "name": "Cover", "role": "Reserve", "start_day": 1, "first_shift": "A3", "is_cover": True}
        ]
    }


def test_valid_strict_solution():
    payload = base_payload()
    response = solve_schedule(ScheduleRequest(**payload))
    assert response.is_valid is True
    assert response.used_best_effort is False
    assert response.used_cover_employee is False
    assert len(response.hard_violations) == 0


def test_best_effort_used_when_strict_has_no_solution():
    payload = base_payload()
    payload["employees"] = payload["employees"][:4] + [payload["employees"][-1]]
    response = solve_schedule(ScheduleRequest(**payload))
    assert response.used_best_effort is True
    assert response.is_valid is True


def test_cover_employee_is_used_when_needed():
    payload = base_payload()
    payload["employees"] = payload["employees"][:4] + [payload["employees"][-1]]
    response = solve_schedule(ScheduleRequest(**payload))
    assert response.used_cover_employee is True
    assert any(stat.is_cover and stat.total_shifts > 0 for stat in response.employee_stats)


def test_start_day_conflict_forces_best_effort_violation():
    payload = base_payload()
    payload["employees"] = [
        {"id": "solo", "name": "Solo", "role": "Operator", "start_day": 2, "first_shift": "A2", "is_cover": False}
    ]
    response = solve_schedule(ScheduleRequest(**payload))
    assert response.used_best_effort is True
    assert response.is_valid is False
    assert any(violation.code == "uncovered_shift" for violation in response.hard_violations)


def test_missing_36h_rest_after_four_day_streak_is_reported():
    payload = deepcopy(base_payload())
    payload["month"] = 2
    payload["year"] = 2026
    payload["employees"] = [
        {"id": "solo", "name": "Solo", "role": "Operator", "start_day": 1, "first_shift": "A2", "is_cover": False}
    ]
    response = solve_schedule(ScheduleRequest(**payload))
    assert response.used_best_effort is True
    assert any(violation.code in {"missing_36h_rest_after_streak", "post_streak_rest_violation", "max_consecutive_days_violation"} for violation in response.hard_violations)


def test_export_files_are_created(tmp_path: Path):
    payload_dict = base_payload()
    payload = ScheduleRequest(**payload_dict)
    response = solve_schedule(payload)
    csv_path = export_csv(payload, response, tmp_path / "schedule.csv")
    xlsx_path = export_xlsx(payload, response, tmp_path / "schedule.xlsx")
    assert csv_path.exists()
    assert xlsx_path.exists()
    csv_text = csv_path.read_text(encoding="utf-8-sig")
    assert "»ÏÂ" in csv_text
    workbook = load_workbook(xlsx_path)
    assert "Schedule" in workbook.sheetnames
    assert "Summary" in workbook.sheetnames


def test_api_schedule_endpoint_returns_expected_shape():
    response = client.post("/api/schedule", json=base_payload())
    assert response.status_code == 200
    body = response.json()
    for key in ["schedule", "hard_violations", "soft_violations", "employee_stats", "score", "is_valid", "used_cover_employee", "used_best_effort", "explanation"]:
        assert key in body
